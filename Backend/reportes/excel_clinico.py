# -*- coding: utf-8 -*-
"""Helper COMPARTIDO para exportaciones Excel con formato institucional.

Mismo criterio que pdf_clinico: un solo estilo (cabecera azul de marca, texto
en negrita blanca, anchos ajustados, panel congelado) para que cualquier
export del sistema se vea igual.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from django.http import HttpResponse

CONTENT_TYPE_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
AZUL_MARCA = "1890FF"
GRIS_SUAVE = "F5F7FA"


def libro_clinico(titulo_hoja: str):
    """Crea (workbook, worksheet) con la hoja titulada."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]  # límite de Excel
    return wb, ws


def escribir_tabla(ws, encabezados: list[str], filas: list[list], fila_inicio: int = 1) -> None:
    """Tabla con cabecera de marca, zebra y anchos según contenido."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill_cabecera = PatternFill("solid", fgColor=AZUL_MARCA)
    fill_zebra = PatternFill("solid", fgColor=GRIS_SUAVE)
    fuente_cabecera = Font(bold=True, color="FFFFFF")

    for col, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_inicio, column=col, value=encabezado)
        celda.font = fuente_cabecera
        celda.fill = fill_cabecera
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for i, fila in enumerate(filas):
        r = fila_inicio + 1 + i
        for col, valor in enumerate(fila, start=1):
            celda = ws.cell(row=r, column=col, value=valor if valor not in (None, "") else "—")
            if i % 2 == 1:
                celda.fill = fill_zebra

    # Anchos: al menos el encabezado, como mucho 42 caracteres
    for col in range(1, len(encabezados) + 1):
        letra = get_column_letter(col)
        largo = max(
            [len(str(encabezados[col - 1]))]
            + [len(str(f[col - 1])) for f in filas if len(f) >= col],
            default=10,
        )
        ws.column_dimensions[letra].width = min(max(largo + 2, 12), 42)

    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)


def escribir_ficha(ws, pares: list[tuple[str, object]], fila_inicio: int = 1) -> int:
    """Bloque campo→valor. Devuelve la siguiente fila libre."""
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=GRIS_SUAVE)
    for i, (etiqueta, valor) in enumerate(pares):
        r = fila_inicio + i
        c1 = ws.cell(row=r, column=1, value=etiqueta)
        c1.font = Font(bold=True)
        c1.fill = fill
        ws.cell(row=r, column=2, value=str(valor) if valor not in (None, "") else "—")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    return fila_inicio + len(pares) + 1


def respuesta_excel(wb, nombre_archivo: str) -> HttpResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fecha = datetime.now().strftime("%Y%m%d")
    response = HttpResponse(buffer.getvalue(), content_type=CONTENT_TYPE_XLSX)
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}_{fecha}.xlsx"'
    return response
