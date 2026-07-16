"""Views module."""
import logging

from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from core.permissions import FetalMedicalPermission

logger = logging.getLogger("ecografias")
import contextlib
from datetime import timedelta

from django.core.cache import cache
from django.db.models import Count
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend

from embarazos.models import Embarazo
from ia_medica.services.orthanc_service import orthanc_service

from .models import Ecografia, ImagenEcografia
from .serializers import (
    EcografiaCreateUpdateSerializer,
    EcografiaListSerializer,
    EcografiaSerializer,
    ImagenEcografiaSerializer,
)


class EcografiaViewSet(viewsets.ModelViewSet):
    """ViewSet for comprehensive Ultrasound management.

    Manages ultrasound examinations, fetal biometry, anatomy assessment,
    and image uploads.

    **Endpoints:**
    - `GET /api/ecografias/` - List all ultrasounds (paginated)
    - `POST /api/ecografias/` - Create new ultrasound
    - `GET /api/ecografias/{id}/` - Get ultrasound details
    - `PUT /api/ecografias/{id}/` - Update ultrasound (full)
    - `PATCH /api/ecografias/{id}/` - Update ultrasound (partial)
    - `DELETE /api/ecografias/{id}/` - Delete ultrasound
    - `POST /api/ecografias/{id}/subir_imagen/` - Upload image to ultrasound
    - `GET /api/ecografias/{id}/imagenes/` - Get all images for ultrasound
    - `DELETE /api/ecografias/{id}/eliminar_imagen/` - Delete specific image
    - `GET /api/ecografias/estadisticas/` - General statistics
    - `GET /api/ecografias/por_embarazo/embarazo_id=X` - Ultrasounds by pregnancy
    - `GET /api/ecografias/{id}/reporte_completo/` - Complete ultrasound report

    **Authentication:** JWT Bearer token required
    **Permissions:** Authenticated users
    """

    queryset = (
        Ecografia.objects.select_related(
            "embarazo", "embarazo__paciente", "paciente", "medico",
        )
        .prefetch_related("biometria", "anatomia", "anexos", "imagenes")
        .all()
    )

    permission_classes = [FetalMedicalPermission]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    filterset_fields = [
        "embarazo",
        "paciente",
        "tipo_ecografia",
        "medico",
        "vitalidad_fetal",
    ]
    search_fields = [
        "paciente__nombre",
        "paciente__apellido_paterno",
        "paciente__id_clinico",
        "diagnostico",
        "observaciones",
    ]
    ordering_fields = ["fecha_ecografia", "edad_gestacional_semanas", "fecha_registro"]

    def get_serializer_class(self):
        """Retorna el serializer apropiado según la acción"""
        if self.action == "list":
            return EcografiaListSerializer
        if self.action in ["create", "update", "partial_update"]:
            return EcografiaCreateUpdateSerializer
        return EcografiaSerializer

    def get_queryset(self):
        """Filtra el queryset según parámetros avanzados"""
        queryset = super().get_queryset()

        # Filtrar por rango de fechas
        fecha_desde = self.request.query_params.get("fecha_desde", None)
        fecha_hasta = self.request.query_params.get("fecha_hasta", None)

        if fecha_desde:
            with contextlib.suppress(Exception):
                queryset = queryset.filter(fecha_ecografia__gte=fecha_desde)

        if fecha_hasta:
            with contextlib.suppress(Exception):
                queryset = queryset.filter(fecha_ecografia__lte=fecha_hasta)

        # Filtrar por edad gestacional
        semanas_min = self.request.query_params.get("semanas_min", None)
        semanas_max = self.request.query_params.get("semanas_max", None)

        if semanas_min:
            with contextlib.suppress(ValueError, TypeError):
                queryset = queryset.filter(
                    edad_gestacional_semanas__gte=int(semanas_min),
                )

        if semanas_max:
            with contextlib.suppress(ValueError, TypeError):
                queryset = queryset.filter(
                    edad_gestacional_semanas__lte=int(semanas_max),
                )

        # Filtrar solo embarazos activos
        solo_activos = self.request.query_params.get("solo_activos", None)
        if solo_activos == "true":
            queryset = queryset.filter(embarazo__estado="activo")

        return queryset.order_by("-fecha_ecografia")

    def perform_create(self, serializer):
        """Asignar automáticamente el médico que está logueado
        Solo si el usuario es médico
        También asignar created_by para trazabilidad
        """
        # ✅ Asignar médico si el usuario es médico
        if getattr(self.request.user, 'rol', '') == "medico":
            serializer.save(
                medico=self.request.user,
                created_by=self.request.user,
                updated_by=self.request.user,
            )
        else:
            serializer.save(created_by=self.request.user, updated_by=self.request.user)

    def create(self, request, *args, **kwargs):
        """Crear nueva ecografía con validaciones"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        # Retornar con serializer completo
        assert serializer.instance is not None
        ecografia = (
            Ecografia.objects.select_related(
                "embarazo", "embarazo__paciente", "paciente", "medico",
            )
            .prefetch_related("biometria", "anatomia", "anexos")
            .get(id=serializer.instance.id)
        )

        return_serializer = EcografiaSerializer(ecografia)
        headers = self.get_success_headers(return_serializer.data)
        return Response(
            return_serializer.data, status=status.HTTP_201_CREATED, headers=headers,
        )

    def perform_update(self, serializer):
        """Asignar updated_by para trazabilidad"""
        serializer.save(updated_by=self.request.user)

    def update(self, request, *args, **kwargs):
        """Actualizar ecografía"""
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        # Recargar con relaciones
        ecografia = (
            Ecografia.objects.select_related(
                "embarazo", "embarazo__paciente", "paciente", "medico",
            )
            .prefetch_related("biometria", "anatomia", "anexos")
            .get(id=instance.id)
        )

        return_serializer = EcografiaSerializer(ecografia)
        return Response(return_serializer.data)

    def destroy(self, request, *args, **kwargs):
        """Eliminar ecografía"""
        instance = self.get_object()
        ecografia_id = instance.id
        tipo = instance.tipo_ecografia
        fecha = instance.fecha_ecografia

        self.perform_destroy(instance)

        return Response(
            {
                "message": "Ecografía eliminada correctamente",
                "ecografia_id": ecografia_id,
                "tipo": tipo,
                "fecha": str(fecha),
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], parser_classes=[MultiPartParser, FormParser])
    def subir_imagen(self, request, _pk=None):
        """Endpoint para subir imágenes a una ecografía

        Parámetros:
            pass
        - imagen: archivo de imagen (requerido)
        - tipo_imagen: tipo de imagen (opcional)
        - titulo: título de la imagen (opcional)
        - descripcion: descripción (opcional)
        - orden: orden de visualización (opcional)
        - es_imagen_principal: booleano (opcional)
        """
        ecografia = self.get_object()

        # ✅ CORREGIDO: usar 'imagen' en lugar de 'archivo'
        imagen_archivo = request.FILES.get("imagen")

        if not imagen_archivo:
            return Response(
                {"error": "No se proporcionó ningún archivo de imagen"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar tipo de archivo — gif excluido (polyglot attack vector, sin valor médico)
        allowed_extensions = ["jpg", "jpeg", "png", "bmp", "tiff", "dcm"]
        file_extension = imagen_archivo.name.rsplit(".", 1)[-1].lower()

        if file_extension not in allowed_extensions:
            return Response(
                {
                    "error": f"Formato no permitido. Formatos médicos aceptados: {', '.join(allowed_extensions)}",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ✅ Crear imagen con trazabilidad
        imagen = ImagenEcografia.objects.create(
            ecografia=ecografia,
            imagen=imagen_archivo,
            tipo_imagen=request.data.get("tipo_imagen", "general"),
            titulo=request.data.get("titulo", "Imagen sin título"),
            descripcion=request.data.get("descripcion", ""),
            orden=int(request.data.get("orden", 1)),
            es_imagen_principal=request.data.get("es_imagen_principal", "false").lower()
            == "true",
            created_by=request.user,
            updated_by=request.user,
        )

        # Subir a Orthanc (PACS) solo para archivos DICOM. Best-effort: si
        # Orthanc no esta disponible, la imagen ya quedo guardada localmente
        # arriba y el usuario no debe ver un error por esto — se degrada
        # igual que el analisis de IA automatico (ver bloque siguiente).
        if file_extension == "dcm":
            try:
                resultado_orthanc = orthanc_service.subir_dicom(imagen.imagen.path)
                if resultado_orthanc.get("status") == "success":
                    imagen.orthanc_instance_id = resultado_orthanc.get("orthanc_id")
                    imagen.orthanc_study_id = resultado_orthanc.get("parent_study")
                    imagen.save(update_fields=["orthanc_instance_id", "orthanc_study_id"])
                else:
                    logger.warning(
                        "No se pudo subir imagen %s a Orthanc: %s",
                        imagen.id, resultado_orthanc.get("message"),
                    )
            except Exception as _e:
                logger.warning(
                    "Error subiendo imagen %s a Orthanc (PACS no disponible?): %s",
                    imagen.id, _e,
                )

        # Disparar análisis IA automáticamente al subir imagen
        try:
            imagen.analizar_con_ia()
        except Exception as _e:
            logger.warning(
                "Análisis IA automático falló para imagen %s: %s",
                getattr(imagen, 'id', None),
                _e,
            )

        serializer = ImagenEcografiaSerializer(imagen, context={"request": request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"])
    def imagenes(self, request, _pk=None):
        """Obtener todas las imágenes de una ecografía"""
        ecografia = self.get_object()
        imagenes = ecografia.imagenes.all().order_by("orden")
        serializer = ImagenEcografiaSerializer(
            imagenes, many=True, context={"request": request},
        )
        return Response({"total": imagenes.count(), "imagenes": serializer.data})

    @action(detail=True, methods=["delete"])
    def eliminar_imagen(self, request, _pk=None):
        """Eliminar una imagen específica"""
        ecografia = self.get_object()
        imagen_id = request.data.get("imagen_id")

        if not imagen_id:
            return Response(
                {"error": "Se requiere imagen_id"}, status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            imagen = ImagenEcografia.objects.get(id=imagen_id, ecografia=ecografia)

            # Eliminar archivo físico
            if imagen.imagen:
                imagen.imagen.delete(save=False)

            imagen.delete()

            return Response(
                {"mensaje": "Imagen eliminada correctamente", "imagen_id": imagen_id},
            )
        except ImagenEcografia.DoesNotExist:
            return Response(
                {"error": "Imagen no encontrada"}, status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=False, methods=["get"])
    def estadisticas(self, _request):
        """Obtener estadisticas generales de ecografias. Cached for 60 seconds."""
        # Clave global (no por usuario): las estadísticas son idénticas para todos.
        # Clave por usuario generaba N×8 queries simultáneas con N usuarios activos.
        cache_key = "ecografia_stats:globales"
        cached_result = cache.get(cache_key)
        if cached_result is not None:
            return Response(cached_result)

        total = Ecografia.objects.count()

        # Estadisticas de la ultima semana
        hace_una_semana = timezone.localdate() - timedelta(days=7)
        ecografias_semana = Ecografia.objects.filter(
            fecha_ecografia__gte=hace_una_semana,
        ).count()

        # Estadisticas del ultimo mes
        hace_un_mes = timezone.localdate() - timedelta(days=30)
        ecografias_mes = Ecografia.objects.filter(
            fecha_ecografia__gte=hace_un_mes,
        ).count()

        # Por tipo
        por_tipo = {}
        for tipo_code, tipo_nombre in Ecografia.TIPO_CHOICES:
            count = Ecografia.objects.filter(tipo_ecografia=tipo_code).count()
            por_tipo[tipo_nombre] = count

        # Con/sin vitalidad fetal
        con_vitalidad = Ecografia.objects.filter(vitalidad_fetal=True).count()
        sin_vitalidad = Ecografia.objects.filter(vitalidad_fetal=False).count()

        # Con imagenes
        con_imagenes = (
            Ecografia.objects.annotate(num_imagenes=Count("imagenes"))
            .filter(num_imagenes__gt=0)
            .count()
        )

        # Por trimestre
        por_trimestre = {
            "primer_trimestre": Ecografia.objects.filter(
                edad_gestacional_semanas__lte=13,
            ).count(),
            "segundo_trimestre": Ecografia.objects.filter(
                edad_gestacional_semanas__gt=13, edad_gestacional_semanas__lte=27,
            ).count(),
            "tercer_trimestre": Ecografia.objects.filter(
                edad_gestacional_semanas__gt=27,
            ).count(),
        }

        result = {
            "total_ecografias": total,
            "ecografias_ultima_semana": ecografias_semana,
            "ecografias_ultimo_mes": ecografias_mes,
            "por_tipo": por_tipo,
            "vitalidad": {
                "con_vitalidad": con_vitalidad,
                "sin_vitalidad": sin_vitalidad,
            },
            "con_imagenes": con_imagenes,
            "distribucion_trimestres": por_trimestre,
        }

        # Cache for 60 seconds
        cache.set(cache_key, result, timeout=60)
        return Response(result)

    @action(detail=False, methods=["get"], url_path="por_embarazo")
    def por_embarazo(self, request):
        """Obtener todas las ecografías de un embarazo específico"""
        embarazo_id = request.query_params.get("embarazo_id", None)

        if not embarazo_id:
            return Response(
                {"error": "Debe proporcionar el parámetro embarazo_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            embarazo = Embarazo.objects.select_related("paciente").get(id=embarazo_id)
        except Embarazo.DoesNotExist:
            return Response(
                {"error": "Embarazo no encontrado"}, status=status.HTTP_404_NOT_FOUND,
            )

        assert self.queryset is not None
        ecografias = self.queryset.filter(embarazo=embarazo).order_by("fecha_ecografia")
        serializer = EcografiaSerializer(
            ecografias, many=True, context={"request": request},
        )

        primera = ecografias.first()
        ultima = ecografias.last()
        return Response(
            {
                "embarazo": {
                    "id": embarazo.id,
                    "paciente": embarazo.paciente.nombre_completo,
                    "numero_gesta": embarazo.numero_gesta,
                    "estado": embarazo.estado,
                },
                "total_ecografias": ecografias.count(),
                "primera_ecografia": primera.fecha_ecografia
                if primera is not None
                else None,
                "ultima_ecografia": ultima.fecha_ecografia
                if ultima is not None
                else None,
                "ecografias": serializer.data,
            },
        )

    @action(detail=True, methods=["get"])
    def reporte_completo(self, _request, _pk=None):
        """Generar reporte completo de una ecografía"""
        ecografia = self.get_object()

        reporte = {
            "ecografia": {
                "id": ecografia.id,
                "fecha": ecografia.fecha_ecografia.strftime("%d/%m/%Y"),
                "tipo": getattr(ecografia, 'get_tipo_ecografia_display')(),
                "indicacion": getattr(ecografia, 'get_indicacion_display')(),
                "edad_gestacional": ecografia.get_edad_gestacional_texto(),
            },
            "paciente": {
                "nombre": ecografia.paciente.nombre_completo,
                "id_clinico": ecografia.paciente.id_clinico,
            },
            "medico": {
                "nombre": ecografia.medico.nombre_completo
                if ecografia.medico
                else None,
            },
            "evaluacion_general": {
                "numero_fetos": ecografia.numero_fetos,
                "vitalidad_fetal": ecografia.vitalidad_fetal,
                "fcf": ecografia.frecuencia_cardiaca_fetal,
                "evaluacion_fcf": ecografia.get_evaluacion_fcf(),
            },
            "liquido_amniotico": {
                "ila": float(ecografia.indice_liquido_amniotico)
                if ecografia.indice_liquido_amniotico
                else None,
                "percentil": ecografia.get_percentil_ila(),
                "estado": ecografia.get_estado_liquido_amniotico(),
            },
            "placenta": {
                "localizacion": ecografia.localizacion_placenta,
                "grado_madurez": ecografia.grado_madurez_placenta,
            },
            "biometria": None,
            "anatomia": None,
            "anexos": None,
            "diagnostico": ecografia.diagnostico,
            "observaciones": ecografia.observaciones,
        }

        # Agregar biometría si existe
        if hasattr(ecografia, "biometria"):
            biometria = ecografia.biometria
            reporte["biometria"] = {
                "dbp": float(biometria.diametro_biparietal)
                if biometria.diametro_biparietal
                else None,
                "cc": float(biometria.circunferencia_cefalica)
                if biometria.circunferencia_cefalica
                else None,
                "ca": float(biometria.circunferencia_abdominal)
                if biometria.circunferencia_abdominal
                else None,
                "lf": float(biometria.longitud_femur)
                if biometria.longitud_femur
                else None,
                "peso_estimado": biometria.peso_fetal_estimado,
                "percentil_peso": biometria.percentil_peso,
                "evaluacion_crecimiento": biometria.get_evaluacion_crecimiento(),
            }

        # Agregar anatomía si existe
        if hasattr(ecografia, "anatomia"):
            anatomia = ecografia.anatomia
            reporte["anatomia"] = {
                "evaluacion_general": anatomia.get_evaluacion_anatomica(),
                "riesgo_cromosomopatias": anatomia.get_riesgo_cromosomopatias(),
                "hallazgos_anormales": anatomia.hallazgos_anormales,
            }

        # Agregar anexos si existe
        if hasattr(ecografia, "anexos"):
            anexos = ecografia.anexos
            reporte["anexos"] = {
                "placenta": anexos.placenta_localizacion,
                "cordon": anexos.get_evaluacion_cordon(),
                "cervix": anexos.get_evaluacion_cervix(),
            }

        return Response(reporte)

    @action(detail=True, methods=["get"])
    def mediciones(self, _request, _pk=None):
        """Obtener todas las mediciones (biometría) de una ecografía"""
        ecografia = self.get_object()
        if hasattr(ecografia, "biometria") and ecografia.biometria:
            from .serializers import BiometriaFetalSerializer
            serializer = BiometriaFetalSerializer(ecografia.biometria)
            return Response(serializer.data)
        return Response(
            {"error": "No hay mediciones de biometría para esta ecografía"},
            status=status.HTTP_404_NOT_FOUND,
        )

    @action(detail=False, methods=["post"])
    def evaluar_icc(self, request):
        """Evaluar Índice Cardiocefálico (ICC)"""
        dbp = request.data.get("dbp")
        cc = request.data.get("cc")
        if not dbp or not cc:
            return Response(
                {"error": "Se requieren dbp (diámetro biparietal) y cc (circunferencia cefálica)"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            dbp = float(dbp)
            cc = float(cc)
            icc = dbp / cc if cc else 0
            return Response({
                "icc": round(icc, 4),
                "evaluacion": "normal" if 0.28 <= icc <= 0.34 else "anormal",
                "interpretacion": (
                    "ICC dentro de rango normal (0.28-0.34)"
                    if 0.28 <= icc <= 0.34
                    else "ICC fuera de rango normal — evaluar dolicocefalia o braquicefalia"
                ),
            })
        except (ValueError, TypeError):
            return Response(
                {"error": "dbp y cc deben ser valores numéricos"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"])
    def evaluar_capurro(self, request):
        """Evaluar edad gestacional por método de Capurro"""
        datos = request.data
        try:
            from .capurro import evaluar_capurro as capurro_func
            resultado = capurro_func(datos)
        except ImportError:
            resultado = {
                "semanas": None,
                "error": "Módulo Capurro no disponible",
                "metodo": "no_disponible",
            }
        except Exception as e:
            resultado = {
                "semanas": None,
                "error": str(e),
                "metodo": "error",
            }
        return Response(resultado)

    @action(detail=True, methods=["get"], url_path="exportar-pdf")
    def exportar_pdf(self, request, pk=None):
        """Exporta ecografía a PDF
        GET /api/ecografias/{id}/exportar-pdf/
        """
        ecografia = self.get_object()
        try:
            import pdfkit
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            html = render_to_string("ecografias/reporte_ecografia_pdf.html", {
                "ecografia": ecografia,
                "paciente": ecografia.paciente,
            })
            pdf = pdfkit.from_string(html, False)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="ecografia_{ecografia.id}.pdf"'
            return response
        except ImportError:
            serializer = self.get_serializer(ecografia)
            return Response({
                "mensaje": "PDF no disponible. Instale pdfkit: pip install pdfkit",
                "ecografia": serializer.data,
            })
        except Exception as e:
            return Response({"error": f"Error generando PDF: {e}"}, status=500)

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Exporta lista de ecografías a Excel
        GET /api/ecografias/exportar-excel/
        """
        try:
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ecografias"
            headers = ["ID", "Paciente", "Fecha", "Tipo", "Edad Gestacional", "Diagnóstico"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            queryset = self.filter_queryset(self.get_queryset())
            for row, eco in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=eco.id)
                ws.cell(row=row, column=2, value=str(eco.paciente))
                ws.cell(row=row, column=3, value=str(eco.fecha_ecografia))
                ws.cell(row=row, column=4, value=eco.tipo_ecografia or "")
                ws.cell(row=row, column=5, value=eco.edad_gestacional_semanas or "")
                ws.cell(row=row, column=6, value=eco.diagnostico or "")
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="ecografias.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response({"mensaje": "Excel no disponible. Instale openpyxl", "total": queryset.count(), "datos": serializer.data})
        except Exception as e:
            return Response({"error": f"Error generando Excel: {e}"}, status=500)
