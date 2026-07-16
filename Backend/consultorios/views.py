"""CONSULTORIOS - VIEWS"""

from datetime import timedelta

from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from core.permissions import EsSoloMedico

from .models import (
    Consultorio,
    HorarioConsultorio,
    MantenimientoConsultorio,
    OcupacionConsultorio,
    ReservaConsultorio,
)
from .serializers import (
    ConsultorioSerializer,
    HorarioConsultorioSerializer,
    MantenimientoConsultorioSerializer,
    OcupacionConsultorioSerializer,
    ReservaConsultorioSerializer,
)


class ConsultorioViewSet(viewsets.ModelViewSet):
    """Consultorioviewset"""
    queryset = Consultorio.objects.all()
    serializer_class = ConsultorioSerializer
    permission_classes = [IsAuthenticated, EsSoloMedico]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["activo", "en_mantenimiento", "tipo_consultorio"]
    search_fields = ["nombre", "codigo", "ubicacion"]
    ordering = ["nombre"]

    def perform_create(self, serializer):
        """✅ TRAZABILIDAD: Auto-asignar created_by al crear"""
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        """✅ TRAZABILIDAD: Auto-asignar updated_by al actualizar"""
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=["get"])
    def disponibles(self, _request):
        """Disponibles"""
        consultorios = self.get_queryset().filter(activo=True, en_mantenimiento=False)
        serializer = self.get_serializer(consultorios, many=True)
        return Response(serializer.data)

    # ─────────────────────────────────────────────────────────────────────
    # DISPONIBILIDAD
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=["get"])
    def verificar_disponibilidad(self, request, pk=None):
        """Verifica si un consultorio está disponible en fecha y hora"""
        consultorio = self.get_object()
        fecha = request.query_params.get("fecha")
        hora_inicio = request.query_params.get("hora_inicio")
        hora_fin = request.query_params.get("hora_fin")

        if not all([fecha, hora_inicio, hora_fin]):
            return Response(
                {
                    "error": (
                        "Debe proporcionar los parámetros "
                        "fecha, hora_inicio y hora_fin"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        disponible = True
        motivo = None

        if not consultorio.esta_disponible():
            disponible = False
            motivo = "El consultorio no está disponible actualmente"

        if disponible:
            from datetime import datetime as dt
            fecha_inicio_dt = dt.strptime(f"{fecha} {hora_inicio}", "%Y-%m-%d %H:%M")
            fecha_fin_dt = dt.strptime(f"{fecha} {hora_fin}", "%Y-%m-%d %H:%M")

            ocupada = OcupacionConsultorio.objects.filter(
                consultorio=consultorio,
                estado__in=["programada", "en_curso"],
                fecha_inicio__lt=fecha_fin_dt,
                fecha_fin__gt=fecha_inicio_dt,
            ).exists()

            if ocupada:
                disponible = False
                motivo = "El consultorio ya está ocupado en ese horario"

        if disponible:
            reservada = ReservaConsultorio.objects.filter(
                consultorio=consultorio,
                estado__in=["pendiente", "aprobada"],
                fecha_reserva=fecha,
                hora_inicio__lt=hora_fin,
                hora_fin__gt=hora_inicio,
            ).exists()

            if reservada:
                disponible = False
                motivo = "El consultorio tiene una reserva en ese horario"

        return Response({"disponible": disponible, "motivo": motivo})

    # ─────────────────────────────────────────────────────────────────────
    # HORARIOS
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=["get", "post"], url_path="horarios")
    def horarios(self, request, pk=None):
        """Lista o crea horarios para un consultorio"""
        consultorio = self.get_object()

        if request.method == "GET":
            horarios = HorarioConsultorio.objects.filter(consultorio=consultorio)
            serializer = HorarioConsultorioSerializer(horarios, many=True)
            return Response(serializer.data)

        serializer = HorarioConsultorioSerializer(
            data={**request.data, "consultorio": consultorio.id}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=True,
        methods=["patch", "delete"],
        url_path=r"horarios/(?P<horario_id>\d+)",
    )
    def horario_detail(self, request, pk=None, horario_id=None):
        """Actualiza o elimina un horario específico"""
        consultorio = self.get_object()
        try:
            horario = HorarioConsultorio.objects.get(
                consultorio=consultorio, id=horario_id
            )
        except HorarioConsultorio.DoesNotExist:
            return Response(
                {"error": "Horario no encontrado"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if request.method == "PATCH":
            serializer = HorarioConsultorioSerializer(
                horario, data=request.data, partial=True
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)

        horario.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ─────────────────────────────────────────────────────────────────────
    # OCUPACIONES (detail)
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=["get"])
    def ocupaciones(self, request, pk=None):
        """Lista ocupaciones de un consultorio"""
        consultorio = self.get_object()
        qs = OcupacionConsultorio.objects.filter(consultorio=consultorio)

        fecha = request.query_params.get("fecha")
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")
        estado = request.query_params.get("estado")

        if fecha:
            qs = qs.filter(fecha_inicio__date=fecha)
        if fecha_inicio:
            qs = qs.filter(fecha_inicio__date__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(fecha_inicio__date__lte=fecha_fin)
        if estado:
            qs = qs.filter(estado=estado)

        serializer = OcupacionConsultorioSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def ocupacion_actual(self, request, pk=None):
        """Obtiene la ocupación actual de un consultorio"""
        consultorio = self.get_object()
        ocupacion = (
            OcupacionConsultorio.objects.filter(
                consultorio=consultorio,
                estado__in=["programada", "en_curso"],
            )
            .order_by("-fecha_inicio")
            .first()
        )

        if not ocupacion:
            return Response(
                {"detail": "No hay ocupación actual para este consultorio"},
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = OcupacionConsultorioSerializer(ocupacion)
        return Response(serializer.data)

    # ─────────────────────────────────────────────────────────────────────
    # OCUPACIONES (global)
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["post"], url_path="ocupaciones")
    def registrar_ocupacion(self, request):
        """Registra una nueva ocupación"""
        serializer = OcupacionConsultorioSerializer(
            data={**request.data, "created_by": request.user.id}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"ocupaciones/(?P<ocupacion_id>\d+)/finalizar",
    )
    def finalizar_ocupacion(self, request, ocupacion_id=None):
        """Finaliza una ocupación"""
        try:
            ocupacion = OcupacionConsultorio.objects.get(id=ocupacion_id)
        except OcupacionConsultorio.DoesNotExist:
            return Response(
                {"error": "Ocupación no encontrada"},
                status=status.HTTP_404_NOT_FOUND,
            )

        ocupacion.estado = "finalizada"
        ocupacion.fecha_fin = timezone.now()
        ocupacion.save(update_fields=["estado", "fecha_fin"])

        serializer = OcupacionConsultorioSerializer(ocupacion)
        return Response(serializer.data)

    # ─────────────────────────────────────────────────────────────────────
    # RESERVAS
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="reservas")
    def listar_reservas(self, request):
        """Lista reservas de consultorios"""
        qs = ReservaConsultorio.objects.all()

        consultorio = request.query_params.get("consultorio")
        fecha = request.query_params.get("fecha")
        estado = request.query_params.get("estado")
        medico = request.query_params.get("medico")

        if consultorio:
            qs = qs.filter(consultorio_id=consultorio)
        if fecha:
            qs = qs.filter(fecha_reserva=fecha)
        if estado:
            qs = qs.filter(estado=estado)
        if medico:
            qs = qs.filter(medico_id=medico)

        serializer = ReservaConsultorioSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="reservas")
    def crear_reserva(self, request):
        """Crea una nueva reserva"""
        serializer = ReservaConsultorioSerializer(
            data={**request.data, "created_by": request.user.id}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"reservas/(?P<reserva_id>\d+)/aprobar",
    )
    def aprobar_reserva(self, request, reserva_id=None):
        """Aprueba una reserva pendiente"""
        try:
            reserva = ReservaConsultorio.objects.get(id=reserva_id)
        except ReservaConsultorio.DoesNotExist:
            return Response(
                {"error": "Reserva no encontrada"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if reserva.estado != "pendiente":
            return Response(
                {
                    "error": (
                        f"No se puede aprobar una reserva "
                        f"en estado {reserva.get_estado_display()}"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        reserva.estado = "aprobada"
        reserva.save(update_fields=["estado"])

        serializer = ReservaConsultorioSerializer(reserva)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"reservas/(?P<reserva_id>\d+)/rechazar",
    )
    def rechazar_reserva(self, request, reserva_id=None):
        """Rechaza una reserva pendiente"""
        try:
            reserva = ReservaConsultorio.objects.get(id=reserva_id)
        except ReservaConsultorio.DoesNotExist:
            return Response(
                {"error": "Reserva no encontrada"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if reserva.estado != "pendiente":
            return Response(
                {
                    "error": (
                        f"No se puede rechazar una reserva "
                        f"en estado {reserva.get_estado_display()}"
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        reserva.estado = "rechazada"
        reserva.save(update_fields=["estado"])

        serializer = ReservaConsultorioSerializer(reserva)
        return Response(serializer.data)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"reservas/(?P<reserva_id>\d+)/cancelar",
    )
    def cancelar_reserva(self, request, reserva_id=None):
        """Cancela una reserva"""
        try:
            reserva = ReservaConsultorio.objects.get(id=reserva_id)
        except ReservaConsultorio.DoesNotExist:
            return Response(
                {"error": "Reserva no encontrada"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if reserva.estado == "cancelada":
            return Response(
                {"error": "La reserva ya está cancelada"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        reserva.estado = "cancelada"
        reserva.save(update_fields=["estado"])

        serializer = ReservaConsultorioSerializer(reserva)
        return Response(serializer.data)

    # ─────────────────────────────────────────────────────────────────────
    # MANTENIMIENTOS
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="mantenimientos")
    def listar_mantenimientos(self, request):
        """Lista mantenimientos de consultorios"""
        qs = MantenimientoConsultorio.objects.all()

        consultorio = request.query_params.get("consultorio")
        tipo = request.query_params.get("tipo")
        estado = request.query_params.get("estado")

        if consultorio:
            qs = qs.filter(consultorio_id=consultorio)
        if tipo:
            qs = qs.filter(tipo=tipo)
        if estado:
            qs = qs.filter(estado=estado)

        serializer = MantenimientoConsultorioSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="mantenimientos")
    def programar_mantenimiento(self, request):
        """Programa un nuevo mantenimiento"""
        serializer = MantenimientoConsultorioSerializer(
            data={**request.data, "created_by": request.user.id}
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(
        detail=False,
        methods=["post"],
        url_path=r"mantenimientos/(?P<mantenimiento_id>\d+)/completar",
    )
    def completar_mantenimiento(self, request, mantenimiento_id=None):
        """Marca un mantenimiento como completado"""
        try:
            mantenimiento = MantenimientoConsultorio.objects.get(
                id=mantenimiento_id
            )
        except MantenimientoConsultorio.DoesNotExist:
            return Response(
                {"error": "Mantenimiento no encontrado"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if mantenimiento.estado == "completado":
            return Response(
                {"error": "El mantenimiento ya fue completado"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        mantenimiento.estado = "completado"
        mantenimiento.fecha_realizacion = timezone.now()
        mantenimiento.realizado_por = request.user
        if "trabajo_realizado" in request.data:
            mantenimiento.descripcion = request.data["trabajo_realizado"]

        mantenimiento.save(
            update_fields=[
                "estado",
                "fecha_realizacion",
                "realizado_por",
                "descripcion",
            ]
        )

        serializer = MantenimientoConsultorioSerializer(mantenimiento)
        return Response(serializer.data)

    # ─────────────────────────────────────────────────────────────────────
    # ESTADÍSTICAS
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=True, methods=["get"])
    def estadisticas(self, request, pk=None):
        """Obtiene estadísticas de uso de un consultorio"""
        consultorio = self.get_object()
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")

        hoy = timezone.now().date()
        if not fecha_fin:
            fecha_fin = hoy.isoformat()
        if not fecha_inicio:
            fecha_inicio = (hoy - timedelta(days=30)).isoformat()

        ocupaciones = OcupacionConsultorio.objects.filter(
            consultorio=consultorio,
            fecha_inicio__date__gte=fecha_inicio,
            fecha_inicio__date__lte=fecha_fin,
        )

        completadas = ocupaciones.filter(
            estado__in=["en_curso", "finalizada"]
        )
        total_ocupaciones = ocupaciones.count()

        horas_ocupadas = sum(
            ((o.fecha_fin or timezone.now()) - o.fecha_inicio).total_seconds()
            / 3600
            for o in completadas
        )

        mantenimientos = MantenimientoConsultorio.objects.filter(
            consultorio=consultorio,
            fecha_programada__date__gte=fecha_inicio,
            fecha_programada__date__lte=fecha_fin,
        )

        horas_disponibles = ocupaciones.values("fecha_inicio__date").distinct().count() * 8.0
        horas_ocupadas_r = round(horas_ocupadas, 2)
        tasa = (
            round((horas_ocupadas_r / horas_disponibles) * 100, 2)
            if horas_disponibles > 0
            else 0
        )

        medicos_ids = (
            ocupaciones.exclude(medico=None)
            .values_list("medico", flat=True)
            .distinct()
        )

        response_data = {
            "consultorio_id": consultorio.id,
            "consultorio_nombre": consultorio.nombre,
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "horas_totales_disponibles": round(horas_disponibles, 2),
            "horas_ocupadas": horas_ocupadas_r,
            "horas_libres": round(max(horas_disponibles - horas_ocupadas_r, 0), 2),
            "tasa_ocupacion": tasa,
            "total_ocupaciones": total_ocupaciones,
            "total_procedimientos": 0,
            "total_emergencias": 0,
            "medicos_usuarios": len(medicos_ids),
            "medico_mas_frecuente": None,
            "total_mantenimientos": mantenimientos.count(),
            "horas_mantenimiento": 0,
            "promedio_ocupacion_dia": 0,
            "dia_mas_ocupado": None,
            "hora_pico": None,
        }

        return Response(response_data)

    # ─────────────────────────────────────────────────────────────────────
    # REPORTES
    # ─────────────────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="reporte-uso")
    def reporte_uso(self, request):
        """Genera reporte de uso de consultorios"""
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")
        consultorio_id = request.query_params.get("consultorio")

        hoy = timezone.now().date()
        if not fecha_fin:
            fecha_fin = hoy.isoformat()
        if not fecha_inicio:
            fecha_inicio = (hoy - timedelta(days=30)).isoformat()

        consultorios = Consultorio.objects.all()
        if consultorio_id:
            consultorios = consultorios.filter(id=consultorio_id)

        data = []
        for c in consultorios:
            ocupaciones = OcupacionConsultorio.objects.filter(
                consultorio=c,
                fecha_inicio__date__gte=fecha_inicio,
                fecha_inicio__date__lte=fecha_fin,
            )
            completadas = ocupaciones.filter(
                estado__in=["en_curso", "finalizada"]
            )

            horas = sum(
                (
                    (o.fecha_fin or timezone.now()) - o.fecha_inicio
                ).total_seconds()
                / 3600
                for o in completadas
            )

            data.append({
                "consultorio_id": c.id,
                "consultorio_nombre": c.nombre,
                "total_ocupaciones": ocupaciones.count(),
                "horas_ocupadas": round(horas, 2),
                "tasa_ocupacion": round(
                    min(horas / ((ocupaciones.count() or 1) * 1.0) * 100, 100), 2
                ),
            })

        return Response({
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "total_consultorios": len(data),
            "datos": data,
        })

    @action(detail=False, methods=["get"], url_path="reporte-pdf")
    def reporte_pdf(self, request):
        """Exporta reporte de consultorios como archivo descargable"""
        fecha_inicio = request.query_params.get("fecha_inicio", "")
        fecha_fin = request.query_params.get("fecha_fin", "")

        consultorios = Consultorio.objects.all()
        lines = [
            "REPORTE DE CONSULTORIOS",
            f"Periodo: {fecha_inicio} - {fecha_fin}",
            "",
            f"{'Consultorio':<25} {'Estado':<15} {'Ocupaciones':<12}",
            "-" * 55,
        ]
        for c in consultorios:
            estado = "Disponible" if c.esta_disponible() else "No disponible"
            total_oc = c.ocupaciones.count()
            lines.append(f"{c.nombre:<25} {estado:<15} {total_oc:<12}")

        response = HttpResponse(
            "\n".join(lines), content_type="text/plain; charset=utf-8"
        )
        response["Content-Disposition"] = (
            'attachment; filename="reporte_consultorios.txt"'
        )
        return response

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Exporta lista de consultorios a Excel
        GET /api/consultorios/exportar-excel/
        """
        try:
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Consultorios"
            headers = ["ID", "Nombre", "Ubicación", "Estado", "Capacidad"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            queryset = self.filter_queryset(self.get_queryset())
            for row, c in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=c.id)
                ws.cell(row=row, column=2, value=c.nombre)
                ws.cell(row=row, column=3, value=c.ubicacion or "")
                ws.cell(row=row, column=4, value=c.estado or "")
                ws.cell(row=row, column=5, value=c.capacidad or "")
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = 'attachment; filename="consultorios.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response({"mensaje": "Excel no disponible. Instale openpyxl", "total": queryset.count(), "datos": serializer.data})
        except Exception as e:
            return Response({"error": f"Error generando Excel: {e}"}, status=500)
