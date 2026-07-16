"""=============================================================================
MÓDULO: PARTOS - VIEWS COMPLETO V4.0
=============================================================================
✅ ViewSets COMPLETOS con todas las funcionalidades
✅ Filtros avanzados, búsquedas y ordenamiento
✅ Acciones personalizadas (@action)
✅ Estadísticas detalladas
✅ Select_related y prefetch_related para optimización
✅ Manejo de errores personalizado
=============================================================================
"""

from datetime import timedelta

from django.db.models import Avg, Count, Max, Min, Q
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response

from core.permissions import FetalMedicalPermission

from .errors import (
    PartoAlreadyFinalizadoError,
)
from .models import (
    ApgarScoreDetallado,
    ComplicacionParto,
    Parto,
    PartogramaRegistro,
    RecienNacido,
)
from .serializers import (
    ApgarScoreDetalladoSerializer,
    ComplicacionPartoSerializer,
    PartogramaRegistroSerializer,
    PartoSerializer,
    RecienNacidoResumenSerializer,
    RecienNacidoSerializer,
)

# ═══════════════════════════════════════════════════════════════════════════
# PARTO VIEWSET - COMPLETO
# ═══════════════════════════════════════════════════════════════════════════


class PartoViewSet(viewsets.ModelViewSet):
    """ViewSet COMPLETO para gestionar partos con todas las funcionalidades

    ENDPOINTS:
        pass
    - list: Lista todos los partos (con paginación)
    - create: Crear nuevo parto
    - retrieve: Obtener detalle de parto
    - update/partial_update: Actualizar parto
    - destroy: Eliminar parto

    ACCIONES PERSONALIZADAS:
        pass
    - estadisticas_generales: Estadísticas de todos los partos
    - partos_recientes: Partos de las últimas 24 horas
    - partos_en_curso: Partos actualmente en curso
    - partos_finalizados: Partos finalizados en rango de fechas
    - estadisticas_por_tipo: Estadísticas agrupadas por tipo de parto
    - finalizar_parto: Marcar parto como finalizado
    - reabrir_parto: Reabrir un parto finalizado
    - por_paciente: Partos de un paciente específico
    - por_embarazo: Partos de un embarazo específico
    - con_complicaciones: Partos que tienen complicaciones
    - cesarea_statistics: Estadísticas de cesáreas
    """

    queryset = Parto.objects.all()
    permission_classes = [FetalMedicalPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    # Campos de filtrado
    filterset_fields = {
        "tipo_parto": ["exact", "in"],
        "parto_finalizado": ["exact"],
        "fecha_parto": ["gte", "lte", "date"],
        "presentacion_fetal": ["exact"],
        "hemorragia_postparto": ["exact"],
        "perdida_sanguinea_estimada": ["gte", "lte"],
        "paciente": ["exact"],
        "embarazo": ["exact"],
        "medico_responsable": ["exact"],
    }

    # Campos de búsqueda
    search_fields = [
        "numero_parto",
        "paciente__nombre",
        "paciente__apellido_paterno",
        "paciente__apellido_materno",
        "paciente__id_clinico",
        "observaciones_parto",
    ]

    # Campos de ordenamiento
    ordering_fields = [
        "fecha_parto",
        "fecha_ingreso",
        "numero_parto",
        "perdida_sanguinea_estimada",
        "duracion_trabajo_parto_horas",
    ]
    ordering = ["-fecha_parto"]  # Ordenamiento por defecto

    def get_queryset(self):
        """Optimiza queries con select_related y prefetch_related"""
        queryset = super().get_queryset()

        # Optimización para evitar N+1 queries
        return queryset.select_related(
            "paciente", "embarazo", "medico_responsable", "created_by", "modified_by",
        ).prefetch_related(
            "recien_nacidos",
            "partograma",
            "complicaciones",
        )


    def get_serializer_class(self):
        """Usa serializer completo para lista y detalle para asegurar datos en dashboard"""
        # if self.action == 'list':
        #     return PartoResumenSerializer
        return PartoSerializer

    def perform_create(self, serializer):
        """Guarda el usuario que crea el parto"""
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        """Guarda el usuario que modifica el parto"""
        serializer.save(modified_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """✅ Eliminar parto completamente de la base de datos (hard delete)"""
        # ✅ SEGURIDAD: Solo administradores pueden eliminar
        if getattr(request.user, 'rol', '') != "administrador":
            return Response(
                {
                    "error": "Sin permisos para eliminar",
                    "detalle": "Solo los administradores pueden eliminar partos del sistema.",
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        from django.db import connection

        instance = self.get_object()
        parto_id = instance.id
        numero_parto = instance.numero_parto
        paciente_nombre = None

        if instance.paciente:
            paciente_nombre = f"{instance.paciente.id_clinico} - {instance.paciente.nombre} {instance.paciente.apellido_paterno}"

        # ✅ Eliminar usando SQL directo (hard delete)
        # Primero eliminar registros dependientes, luego el parto
        with connection.cursor() as cursor:
            # 1. Eliminar recién nacidos
            cursor.execute("DELETE FROM recien_nacidos WHERE parto_id = %s", [parto_id])
            # 2. Eliminar registros de partograma
            cursor.execute(
                "DELETE FROM partograma_registros WHERE parto_id = %s", [parto_id],
            )
            # 3. Eliminar complicaciones
            cursor.execute(
                "DELETE FROM complicaciones_parto WHERE parto_id = %s", [parto_id],
            )
            # 4. Finalmente eliminar el parto
            cursor.execute("DELETE FROM partos WHERE id = %s", [parto_id])

        return Response(
            {
                "message": "Parto eliminado completamente",
                "detalles": {
                    "numero_parto": numero_parto,
                    "paciente": paciente_nombre,
                },
                "id": parto_id,
            },
            status=status.HTTP_200_OK,
        )

    # ═══════════════════════════════════════════════════════════════════════
    # ACCIONES PERSONALIZADAS - ESTADÍSTICAS
    # ═══════════════════════════════════════════════════════════════════════

    @action(detail=False, methods=["get"])
    def estadisticas_generales(self, _request):
        """Estadísticas generales de todos los partos
        GET /api/partos/estadisticas_generales/
        """
        # Total de partos
        total_partos = self.get_queryset().count()

        # Partos finalizados vs en curso
        finalizados = self.get_queryset().filter(parto_finalizado=True).count()
        en_curso = total_partos - finalizados

        # Por tipo de parto
        por_tipo = (
            self.get_queryset()
            .values("tipo_parto")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        # Estadísticas de pérdida sanguínea
        stats_sangre = self.get_queryset().aggregate(
            promedio_perdida=Avg("perdida_sanguinea_estimada"),
            max_perdida=Max("perdida_sanguinea_estimada"),
            min_perdida=Min("perdida_sanguinea_estimada"),
            total_hemorragias=Count("id", filter=Q(hemorragia_postparto=True)),
        )

        # Estadísticas de duración
        stats_duracion = self.get_queryset().aggregate(
            promedio_duracion=Avg("duracion_trabajo_parto_horas"),
            max_duracion=Max("duracion_trabajo_parto_horas"),
            min_duracion=Min("duracion_trabajo_parto_horas"),
        )

        # Recién nacidos
        total_recien_nacidos = RecienNacido.objects.filter(
            parto__in=self.get_queryset(),
        ).count()

        # Complicaciones
        total_complicaciones = ComplicacionParto.objects.filter(
            parto__in=self.get_queryset(),
        ).count()

        partos_con_complicaciones = (
            self.get_queryset().filter(complicaciones__isnull=False).distinct().count()
        )

        return Response(
            {
                "resumen": {
                    "total_partos": total_partos,
                    "finalizados": finalizados,
                    "en_curso": en_curso,
                    "tasa_finalizacion": f"{(finalizados / total_partos * 100):.1f}%"
                    if total_partos > 0
                    else "0%",
                },
                "por_tipo": list(por_tipo),
                "perdida_sanguinea": {
                    "promedio_ml": round(stats_sangre["promedio_perdida"], 2)
                    if stats_sangre["promedio_perdida"]
                    else 0,
                    "maxima_ml": stats_sangre["max_perdida"] or 0,
                    "minima_ml": stats_sangre["min_perdida"] or 0,
                    "total_hemorragias": stats_sangre["total_hemorragias"],
                },
                "duracion_trabajo_parto": {
                    "promedio_horas": round(stats_duracion["promedio_duracion"], 2)
                    if stats_duracion["promedio_duracion"]
                    else 0,
                    "maxima_horas": float(stats_duracion["max_duracion"])
                    if stats_duracion["max_duracion"]
                    else 0,
                    "minima_horas": float(stats_duracion["min_duracion"])
                    if stats_duracion["min_duracion"]
                    else 0,
                },
                "recien_nacidos": {
                    "total": total_recien_nacidos,
                    "promedio_por_parto": round(total_recien_nacidos / total_partos, 2)
                    if total_partos > 0
                    else 0,
                },
                "complicaciones": {
                    "total_complicaciones": total_complicaciones,
                    "partos_con_complicaciones": partos_con_complicaciones,
                    "tasa_complicaciones": f"{(partos_con_complicaciones / total_partos * 100):.1f}%"
                    if total_partos > 0
                    else "0%",
                },
            },
        )

    @action(detail=False, methods=["get"])
    def partos_recientes(self, _request):
        """Partos de las últimas 24 horas
        GET /api/partos/partos_recientes/
        """
        hace_24h = timezone.now() - timedelta(hours=24)
        partos = (
            self.get_queryset()
            .filter(fecha_parto__gte=hace_24h)
            .order_by("-fecha_parto")
        )

        serializer = self.get_serializer(partos, many=True)
        return Response({"total": partos.count(), "partos": serializer.data})

    @action(detail=False, methods=["get"])
    def partos_en_curso(self, _request):
        """Partos actualmente en curso (no finalizados)
        GET /api/partos/partos_en_curso/
        """
        partos = (
            self.get_queryset()
            .filter(parto_finalizado=False)
            .order_by("fecha_inicio_trabajo_parto")
        )

        serializer = self.get_serializer(partos, many=True)
        return Response({"total": partos.count(), "partos": serializer.data})

    @action(detail=False, methods=["get"])
    def partos_finalizados(self, request):
        """Partos finalizados en rango de fechas
        GET /api/partos/partos_finalizados/fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD
        """
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")

        queryset = self.get_queryset().filter(parto_finalizado=True)

        if fecha_inicio:
            queryset = queryset.filter(fecha_parto__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_parto__lte=fecha_fin)

        serializer = self.get_serializer(queryset, many=True)
        return Response({"total": queryset.count(), "partos": serializer.data})

    @action(detail=False, methods=["get"])
    def estadisticas_por_tipo(self, _request):
        """Estadísticas agrupadas por tipo de parto
        GET /api/partos/estadisticas_por_tipo/
        """
        stats = []

        for tipo_codigo, tipo_nombre in Parto.TIPO_PARTO_CHOICES:
            partos_tipo = self.get_queryset().filter(tipo_parto=tipo_codigo)

            stats.append(
                {
                    "tipo": tipo_nombre,
                    "codigo": tipo_codigo,
                    "total": partos_tipo.count(),
                    "promedio_perdida_sanguinea": round(
                        partos_tipo.aggregate(Avg("perdida_sanguinea_estimada"))[
                            "perdida_sanguinea_estimada__avg"
                        ]
                        or 0,
                        2,
                    ),
                    "promedio_duracion": round(
                        partos_tipo.aggregate(Avg("duracion_trabajo_parto_horas"))[
                            "duracion_trabajo_parto_horas__avg"
                        ]
                        or 0,
                        2,
                    ),
                    "con_complicaciones": partos_tipo.filter(
                        complicaciones__isnull=False,
                    )
                    .distinct()
                    .count(),
                },
            )

        return Response({"estadisticas_por_tipo": stats})

    @action(detail=False, methods=["get"])
    def cesarea_statistics(self, _request):
        """Estadísticas específicas de cesáreas
        GET /api/partos/cesarea_statistics/
        """
        cesareas = self.get_queryset().filter(
            tipo_parto__in=[
                "cesarea_electiva",
                "cesarea_urgencia",
                "cesarea_emergencia",
            ],
        )

        total_cesareas = cesareas.count()
        total_partos = self.get_queryset().count()

        por_tipo = cesareas.values("tipo_parto").annotate(total=Count("id"))

        # Indicaciones más comunes
        indicaciones_comunes = (
            cesareas.exclude(indicaciones_cesarea="")
            .values("indicaciones_cesarea")
            .annotate(total=Count("id"))
            .order_by("-total")[:10]
        )

        return Response(
            {
                "total_cesareas": total_cesareas,
                "total_partos": total_partos,
                "tasa_cesareas": f"{(total_cesareas / total_partos * 100):.1f}%"
                if total_partos > 0
                else "0%",
                "distribucion_por_tipo": list(por_tipo),
                "indicaciones_mas_comunes": list(indicaciones_comunes),
            },
        )

    # ═══════════════════════════════════════════════════════════════════════
    # ACCIONES PERSONALIZADAS - GESTIÓN DE PARTO
    # ═══════════════════════════════════════════════════════════════════════

    @action(detail=True, methods=["post"])
    def finalizar_parto(self, request, pk=None):
        """Marca un parto como finalizado
        POST /api/partos/{id}/finalizar_parto/
        """
        parto = self.get_object()

        if parto.parto_finalizado:
            raise PartoAlreadyFinalizadoError()

        parto.parto_finalizado = True
        parto.modified_by = request.user

        # Si no tiene fecha_parto, usar fecha actual
        if not parto.fecha_parto:
            parto.fecha_parto = timezone.now()

        parto.save()

        serializer = self.get_serializer(parto)
        return Response(
            {"message": "Parto finalizado exitosamente", "parto": serializer.data},
        )

    @action(detail=True, methods=["post"])
    def reabrir_parto(self, request, pk=None):
        """Reabre un parto finalizado (solo si tiene permisos)
        POST /api/partos/{id}/reabrir_parto/
        """
        parto = self.get_object()

        if not parto.parto_finalizado:
            return Response(
                {"error": "El parto no está finalizado"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        parto.parto_finalizado = False
        parto.modified_by = request.user
        parto.save()

        serializer = self.get_serializer(parto)
        return Response(
            {"message": "Parto reabierto exitosamente", "parto": serializer.data},
        )

    @action(detail=True, methods=["get"])
    def resumen_completo(self, _request, pk=None):
        """Obtiene resumen completo del parto con todas sus relaciones
        GET /api/partos/{id}/resumen_completo/
        """
        parto = self.get_object()

        return Response(
            {
                "parto": PartoSerializer(parto).data,
                "recien_nacidos": RecienNacidoSerializer(
                    parto.recien_nacidos.all(), many=True,
                ).data,
                "partograma": PartogramaRegistroSerializer(
                    parto.partograma.all(), many=True,
                ).data,
                "complicaciones": ComplicacionPartoSerializer(
                    parto.complicaciones.all(), many=True,
                ).data,
                "estadisticas": {
                    "total_recien_nacidos": parto.recien_nacidos.count(),
                    "total_registros_partograma": parto.partograma.count(),
                    "total_complicaciones": parto.complicaciones.count(),
                },
            },
        )

    # ═══════════════════════════════════════════════════════════════════════
    # ACCIONES PERSONALIZADAS - FILTROS ESPECÍFICOS
    # ═══════════════════════════════════════════════════════════════════════

    @action(detail=False, methods=["get"])
    def por_paciente(self, request):
        """Obtiene todos los partos de un paciente
        GET /api/partos/por_paciente/paciente_id=123
        """
        paciente_id = request.query_params.get("paciente_id")

        if not paciente_id:
            return Response(
                {"error": "Debe proporcionar paciente_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        partos = self.get_queryset().filter(paciente_id=paciente_id)
        serializer = self.get_serializer(partos, many=True)

        return Response({"total": partos.count(), "partos": serializer.data})

    @action(detail=False, methods=["get"])
    def por_embarazo(self, request):
        """Obtiene el parto de un embarazo específico
        GET /api/partos/por_embarazo/embarazo_id=123
        """
        embarazo_id = request.query_params.get("embarazo_id")

        if not embarazo_id:
            return Response(
                {"error": "Debe proporcionar embarazo_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        partos = self.get_queryset().filter(embarazo_id=embarazo_id)
        serializer = self.get_serializer(partos, many=True)

        return Response({"total": partos.count(), "partos": serializer.data})

    @action(detail=False, methods=["get"])
    def con_complicaciones(self, _request):
        """Obtiene partos que tienen complicaciones registradas
        GET /api/partos/con_complicaciones/
        """
        partos = self.get_queryset().filter(complicaciones__isnull=False).distinct()

        serializer = self.get_serializer(partos, many=True)

        return Response({"total": partos.count(), "partos": serializer.data})

    @action(detail=False, methods=["post"], url_path="calcular-apgar")
    def calcular_apgar(self, request):
        """Calcula puntaje Apgar desde el frontend
        POST /api/partos/calcular-apgar/
        Body: {"frecuencia_cardiaca": 2, "esfuerzo_respiratorio": 2, ...}
        """
        try:
            data = request.data
            if isinstance(data, str):
                import json
                data = json.loads(data)
            requeridos = [
                "frecuencia_cardiaca", "esfuerzo_respiratorio",
                "tono_muscular", "reflejos", "color_piel",
            ]
            for campo in requeridos:
                if campo not in data:
                    return Response(
                        {"error": f"Campo requerido: {campo}"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            total = sum(int(data[c]) for c in requeridos)
            if total >= 7:
                clasificacion = "Normal"
            elif total >= 4:
                clasificacion = "Moderadamente deprimido"
            else:
                clasificacion = "Severamente deprimido"
            return Response({
                "score_total": total,
                "clasificacion": clasificacion,
                "componentes": {c: int(data[c]) for c in requeridos},
            })
        except (ValueError, TypeError) as e:
            return Response(
                {"error": f"Datos inválidos: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"], url_path="analizar-paciente")
    def analizar_paciente(self, request):
        """Analiza el historial de partos de una paciente
        GET /api/partos/analizar-paciente/?paciente_id=123
        """
        paciente_id = request.query_params.get("paciente_id")
        if not paciente_id:
            return Response(
                {"error": "Debe proporcionar paciente_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        partos = self.get_queryset().filter(paciente_id=paciente_id)
        total = partos.count()
        if total == 0:
            return Response({
                "paciente_id": int(paciente_id),
                "total_partos": 0,
                "mensaje": "Paciente sin partos registrados",
            })
        partos_data = self.get_serializer(partos, many=True).data
        return Response({
            "paciente_id": int(paciente_id),
            "total_partos": total,
            "partos": partos_data,
        })

    @action(detail=True, methods=["get"], url_path="calcular-riesgo")
    def calcular_riesgo(self, _request, pk=None):
        """Calcula nivel de riesgo de un parto
        GET /api/partos/{id}/calcular-riesgo/
        """
        parto = self.get_object()
        factores = []
        nivel_riesgo = "bajo"
        if parto.tipo_parto and parto.tipo_parto.lower() in ("cesarea", "forceps", "distocico"):
            factores.append("Tipo de parto de alto riesgo")
            nivel_riesgo = "alto"
        if parto.hemorragia_postparto:
            factores.append("Hemorragia registrada")
            nivel_riesgo = "alto" if nivel_riesgo != "alto" else "alto"
        if parto.complicaciones.exists():
            factores.append(f"{parto.complicaciones.count()} complicación(es) registrada(s)")
            if nivel_riesgo == "bajo":
                nivel_riesgo = "medio"
        if parto.edad_gestacional_parto:
            try:
                semanas_parto = int(parto.edad_gestacional_parto.split("+")[0])
                if semanas_parto < 37:
                    factores.append("Parto prematuro")
                    nivel_riesgo = "alto"
            except (ValueError, IndexError):
                pass
        recien_nacidos = parto.recien_nacidos.all()
        for rn in recien_nacidos:
            if rn.peso and rn.peso < 2500:
                factores.append(f"Bajo peso al nacer: {rn.peso}g")
                nivel_riesgo = "alto" if nivel_riesgo != "alto" else "alto"
                break
        return Response({
            "parto_id": parto.id,
            "nivel_riesgo": nivel_riesgo,
            "factores_riesgo": factores,
            "total_factores": len(factores),
        })

    @action(detail=True, methods=["get"], url_path="generar-pdf")
    def generar_pdf(self, _request, pk=None):
        """Genera PDF con resumen del parto
        GET /api/partos/{id}/generar-pdf/
        """
        parto = self.get_object()
        try:
            import pdfkit
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            html = render_to_string("partos/reporte_parto_pdf.html", {
                "parto": parto,
                "paciente": parto.paciente,
                "recien_nacidos": parto.recien_nacidos.all(),
                "complicaciones": parto.complicaciones.all(),
            })
            pdf = pdfkit.from_string(html, False)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="parto_{parto.id}.pdf"'
            return response
        except ImportError:
            return Response({
                "parto_id": parto.id,
                "mensaje": "PDF no disponible. Instale pdfkit: pip install pdfkit",
                "datos": self.get_serializer(parto).data,
            })
        except Exception as e:
            return Response(
                {"error": f"Error generando PDF: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["get"], url_path="partograma-pdf")
    def partograma_pdf(self, _request, pk=None):
        """Genera PDF del partograma
        GET /api/partos/{id}/partograma-pdf/
        """
        parto = self.get_object()
        try:
            import pdfkit
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            registros = parto.partograma.all().order_by("hora_registro")
            html = render_to_string("partos/partograma_pdf.html", {
                "parto": parto,
                "registros": registros,
            })
            pdf = pdfkit.from_string(html, False)
            response = HttpResponse(pdf, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="partograma_{parto.id}.pdf"'
            return response
        except ImportError:
            from .serializers import PartogramaRegistroSerializer
            registros = parto.partograma.all().order_by("hora_registro")
            return Response({
                "parto_id": parto.id,
                "mensaje": "PDF no disponible. Instale pdfkit: pip install pdfkit",
                "registros": PartogramaRegistroSerializer(registros, many=True).data,
            })
        except Exception as e:
            return Response(
                {"error": f"Error generando PDF: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"], url_path="exportar-excel")
    def exportar_excel(self, request):
        """Exporta lista de partos a Excel
        GET /api/partos/exportar-excel/
        """
        try:
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Partos"
            headers = [
                "ID", "Paciente", "Fechahora", "Tipo Parto",
                "Edad Gestacional", "Estado", "Complicaciones",
            ]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            queryset = self.filter_queryset(self.get_queryset())
            for row, parto in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=parto.id)
                ws.cell(row=row, column=2, value=str(parto.paciente))
                ws.cell(row=row, column=3, value=str(parto.fecha_parto))
                ws.cell(row=row, column=4, value=parto.tipo_parto or "")
                ws.cell(row=row, column=5, value=parto.edad_gestacional_parto or "")
                ws.cell(row=row, column=6, value=parto.get_estado_parto())
                ws.cell(row=row, column=7, value=parto.complicaciones.count())
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="partos.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            return Response({
                "mensaje": "Excel no disponible. Instale openpyxl: pip install openpyxl",
                "total": queryset.count(),
                "datos": serializer.data,
            })
        except Exception as e:
            return Response(
                {"error": f"Error exportando Excel: {e}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# ═══════════════════════════════════════════════════════════════════════════
# RECIÉN NACIDO VIEWSET - COMPLETO
# ═══════════════════════════════════════════════════════════════════════════


class RecienNacidoViewSet(viewsets.ModelViewSet):
    """ViewSet COMPLETO para gestionar recién nacidos

    ACCIONES PERSONALIZADAS:
        pass
    - estadisticas_generales: Estadísticas de todos los RN
    - por_peso: RN agrupados por clasificación de peso
    - por_apgar: RN agrupados por score de Apgar
    - con_reanimacion: RN que requirieron reanimación
    - con_malformaciones: RN con malformaciones congénitas
    - por_parto: RN de un parto específico
    """

    queryset = RecienNacido.objects.all()
    permission_classes = [FetalMedicalPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filterset_fields = {
        "parto": ["exact"],
        "sexo": ["exact"],
        "estado_nacimiento": ["exact"],
        "peso_nacimiento": ["gte", "lte"],
        "apgar_5_minutos": ["gte", "lte"],
        "requirio_reanimacion": ["exact"],
        "malformaciones_congenitas": ["exact"],
        "destino_rn": ["exact"],
    }

    search_fields = [
        "parto__numero_parto",
        "parto__paciente__nombre",
        "observaciones_rn",
    ]

    ordering_fields = [
        "fecha_nacimiento",
        "peso_nacimiento",
        "apgar_5_minutos",
        "talla_nacimiento",
    ]
    ordering = ["-fecha_nacimiento"]

    def get_queryset(self):
        """Optimiza queries"""
        return (
            super()
            .get_queryset()
            .select_related("parto", "parto__paciente", "parto__embarazo")
        )

    def get_serializer_class(self):
        """Usa serializer resumido para lista"""
        if self.action == "list":
            return RecienNacidoResumenSerializer
        return RecienNacidoSerializer

    @action(detail=False, methods=["get"])
    def estadisticas_generales(self, _request):
        """Estadísticas generales de recién nacidos"""
        queryset = self.get_queryset()

        return Response(
            {
                "total": queryset.count(),
                "por_sexo": list(queryset.values("sexo").annotate(total=Count("id"))),
                "peso": {
                    "promedio": round(
                        queryset.aggregate(Avg("peso_nacimiento"))[
                            "peso_nacimiento__avg"
                        ]
                        or 0,
                        2,
                    ),
                    "maximo": queryset.aggregate(Max("peso_nacimiento"))[
                        "peso_nacimiento__max"
                    ]
                    or 0,
                    "minimo": queryset.aggregate(Min("peso_nacimiento"))[
                        "peso_nacimiento__min"
                    ]
                    or 0,
                },
                "apgar": {
                    "promedio_1min": round(
                        queryset.aggregate(Avg("apgar_1_minuto"))["apgar_1_minuto__avg"]
                        or 0,
                        2,
                    ),
                    "promedio_5min": round(
                        queryset.aggregate(Avg("apgar_5_minutos"))[
                            "apgar_5_minutos__avg"
                        ]
                        or 0,
                        2,
                    ),
                },
                "reanimacion": {
                    "total_requirieron": queryset.filter(
                        requirio_reanimacion=True,
                    ).count(),
                    "porcentaje": f"{(queryset.filter(requirio_reanimacion=True).count() / queryset.count() * 100):.1f}%"
                    if queryset.count() > 0
                    else "0%",
                },
                "malformaciones": queryset.filter(
                    malformaciones_congenitas=True,
                ).count(),
                "destinos": list(
                    queryset.values("destino_rn").annotate(total=Count("id")),
                ),
            },
        )

    @action(detail=False, methods=["get"])
    def por_peso(self, _request):
        """RN agrupados por clasificación de peso"""
        queryset = self.get_queryset()

        clasificaciones = {
            "extremadamente_bajo": queryset.filter(peso_nacimiento__lt=1000).count(),
            "muy_bajo": queryset.filter(
                peso_nacimiento__gte=1000, peso_nacimiento__lt=1500,
            ).count(),
            "bajo": queryset.filter(
                peso_nacimiento__gte=1500, peso_nacimiento__lt=2500,
            ).count(),
            "normal": queryset.filter(
                peso_nacimiento__gte=2500, peso_nacimiento__lte=4000,
            ).count(),
            "macrosomia_leve": queryset.filter(
                peso_nacimiento__gt=4000, peso_nacimiento__lte=4500,
            ).count(),
            "macrosomia_severa": queryset.filter(peso_nacimiento__gt=4500).count(),
        }

        return Response(clasificaciones)

    @action(detail=False, methods=["get"])
    def por_apgar(self, _request):
        """RN agrupados por score de Apgar a 5 minutos"""
        queryset = self.get_queryset()

        clasificaciones = {
            "critico": queryset.filter(apgar_5_minutos__lt=4).count(),
            "bajo": queryset.filter(
                apgar_5_minutos__gte=4, apgar_5_minutos__lt=6,
            ).count(),
            "moderado": queryset.filter(
                apgar_5_minutos__gte=6, apgar_5_minutos__lt=8,
            ).count(),
            "excelente": queryset.filter(apgar_5_minutos__gte=8).count(),
        }

        return Response(clasificaciones)

    @action(detail=False, methods=["get"])
    def con_reanimacion(self, _request):
        """RN que requirieron reanimación"""
        rns = self.get_queryset().filter(requirio_reanimacion=True)
        serializer = self.get_serializer(rns, many=True)

        return Response({"total": rns.count(), "recien_nacidos": serializer.data})

    @action(detail=False, methods=["get"])
    def con_malformaciones(self, _request):
        """RN con malformaciones congénitas"""
        rns = self.get_queryset().filter(malformaciones_congenitas=True)
        serializer = self.get_serializer(rns, many=True)

        return Response({"total": rns.count(), "recien_nacidos": serializer.data})

    @action(detail=False, methods=["get"])
    def por_parto(self, request):
        """RN de un parto específico"""
        parto_id = request.query_params.get("parto_id")

        if not parto_id:
            return Response(
                {"error": "Debe proporcionar parto_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        rns = self.get_queryset().filter(parto_id=parto_id)
        serializer = self.get_serializer(rns, many=True)

        return Response({"total": rns.count(), "recien_nacidos": serializer.data})

    @action(detail=True, methods=["get"])
    def resumen_completo(self, _request, pk=None):
        """Obtiene resumen completo del recién nacido
        GET /api/partos/recien-nacidos/{id}/resumen_completo/
        """
        rn = self.get_object()

        return Response(
            {
                "recien_nacido": RecienNacidoSerializer(rn).data,
                "apgar_scores": ApgarScoreDetalladoSerializer(
                    rn.apgar_scores.all(), many=True,
                ).data
                if hasattr(rn, "apgar_scores")
                else [],
                "parto_info": {
                    "numero_parto": rn.parto.numero_parto if rn.parto else None,
                    "tipo_parto": getattr(rn.parto, 'get_tipo_parto_display')()
                    if rn.parto
                    else None,
                },
            },
        )


# ═══════════════════════════════════════════════════════════════════════════
# PARTOGRAMA REGISTRO VIEWSET - COMPLETO
# ═══════════════════════════════════════════════════════════════════════════


class PartogramaRegistroViewSet(viewsets.ModelViewSet):
    """ViewSet COMPLETO para gestionar registros de partograma

    ACCIONES PERSONALIZADAS:
        pass
    - por_parto: Registros de un parto específico
    - con_alertas: Registros con alertas activas
    - fcf_anormal: Registros con FCF anormal
    - progreso_lento: Registros con progreso lento
    - grafica_dilatacion: Datos para gráfica de dilatación
    - grafica_fcf: Datos para gráfica de FCF
    """

    queryset = PartogramaRegistro.objects.all()
    serializer_class = PartogramaRegistroSerializer
    permission_classes = [FetalMedicalPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filterset_fields = {
        "parto": ["exact"],
        "dilatacion_cervical": ["gte", "lte"],
        "fcf_baseline": ["gte", "lte"],
        "alerta_fcf_anormal": ["exact"],
        "alerta_progreso_lento": ["exact"],
        "alerta_signos_vitales": ["exact"],
    }

    ordering_fields = ["hora_registro", "horas_trabajo_parto", "dilatacion_cervical"]
    ordering = ["hora_registro"]

    def get_queryset(self):
        """Optimiza queries"""
        return super().get_queryset().select_related("parto")

    @action(detail=False, methods=["get"])
    def por_parto(self, request):
        """Registros de un parto específico"""
        parto_id = request.query_params.get("parto_id")

        if not parto_id:
            return Response(
                {"error": "Debe proporcionar parto_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = (
            self.get_queryset().filter(parto_id=parto_id).order_by("hora_registro")
        )
        serializer = self.get_serializer(registros, many=True)

        return Response({"total": registros.count(), "registros": serializer.data})

    @action(detail=False, methods=["get"])
    def con_alertas(self, _request):
        """Registros con cualquier alerta activa"""
        registros = self.get_queryset().filter(
            Q(alerta_fcf_anormal=True)
            | Q(alerta_progreso_lento=True)
            | Q(alerta_signos_vitales=True),
        )

        serializer = self.get_serializer(registros, many=True)

        return Response({"total": registros.count(), "registros": serializer.data})

    @action(detail=False, methods=["get"])
    def fcf_anormal(self, _request):
        """Registros con FCF anormal"""
        registros = self.get_queryset().filter(alerta_fcf_anormal=True)
        serializer = self.get_serializer(registros, many=True)

        return Response({"total": registros.count(), "registros": serializer.data})

    @action(detail=False, methods=["get"])
    def progreso_lento(self, _request):
        """Registros con progreso lento"""
        registros = self.get_queryset().filter(alerta_progreso_lento=True)
        serializer = self.get_serializer(registros, many=True)

        return Response({"total": registros.count(), "registros": serializer.data})

    @action(detail=False, methods=["get"])
    def grafica_dilatacion(self, request):
        """Datos para gráfica de dilatación cervical"""
        parto_id = request.query_params.get("parto_id")

        if not parto_id:
            return Response(
                {"error": "Debe proporcionar parto_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = (
            self.get_queryset()
            .filter(parto_id=parto_id)
            .order_by("hora_registro")
            .values("hora_registro", "horas_trabajo_parto", "dilatacion_cervical")
        )

        return Response({"datos": list(registros)})

    @action(detail=False, methods=["get"])
    def grafica_fcf(self, request):
        """Datos para gráfica de FCF"""
        parto_id = request.query_params.get("parto_id")

        if not parto_id:
            return Response(
                {"error": "Debe proporcionar parto_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        registros = (
            self.get_queryset()
            .filter(parto_id=parto_id, fcf_baseline__isnull=False)
            .order_by("hora_registro")
            .values("hora_registro", "horas_trabajo_parto", "fcf_baseline")
        )

        return Response({"datos": list(registros)})


# ═══════════════════════════════════════════════════════════════════════════
# COMPLICACIÓN PARTO VIEWSET - COMPLETO
# ═══════════════════════════════════════════════════════════════════════════


class ComplicacionPartoViewSet(viewsets.ModelViewSet):
    """ViewSet COMPLETO para gestionar complicaciones del parto

    ACCIONES PERSONALIZADAS:
        pass
    - por_parto: Complicaciones de un parto específico
    - por_severidad: Complicaciones agrupadas por severidad
    - por_tipo: Complicaciones agrupadas por tipo
    - criticas: Complicaciones críticas
    - requirieron_cirugia: Complicaciones que requirieron cirugía
    - estadisticas: Estadísticas de complicaciones
    """

    queryset = ComplicacionParto.objects.all()
    serializer_class = ComplicacionPartoSerializer
    permission_classes = [FetalMedicalPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filterset_fields = {
        "parto": ["exact"],
        "tipo_complicacion": ["exact", "in"],
        "severidad": ["exact", "in"],
        "requirio_cirugia": ["exact"],
        "resolucion_complicacion": ["exact"],
    }

    search_fields = [
        "descripcion_detallada",
        "tratamiento_realizado",
        "tipo_complicacion",
    ]

    ordering_fields = ["momento_deteccion", "severidad"]
    ordering = ["-momento_deteccion"]

    def get_queryset(self):
        """Optimiza queries"""
        return super().get_queryset().select_related("parto", "medico_responsable")

    @action(detail=False, methods=["get"])
    def por_parto(self, request):
        """Complicaciones de un parto específico"""
        parto_id = request.query_params.get("parto_id")

        if not parto_id:
            return Response(
                {"error": "Debe proporcionar parto_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        complicaciones = self.get_queryset().filter(parto_id=parto_id)
        serializer = self.get_serializer(complicaciones, many=True)

        return Response(
            {"total": complicaciones.count(), "complicaciones": serializer.data},
        )

    @action(detail=False, methods=["get"])
    def por_severidad(self, _request):
        """Complicaciones agrupadas por severidad"""
        stats = (
            self.get_queryset()
            .values("severidad")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        return Response({"distribucion": list(stats)})

    @action(detail=False, methods=["get"])
    def por_tipo(self, _request):
        """Complicaciones agrupadas por tipo"""
        stats = (
            self.get_queryset()
            .values("tipo_complicacion")
            .annotate(total=Count("id"))
            .order_by("-total")
        )

        return Response({"distribucion": list(stats)})

    @action(detail=False, methods=["get"])
    def criticas(self, _request):
        """Complicaciones críticas"""
        complicaciones = self.get_queryset().filter(severidad="critica")
        serializer = self.get_serializer(complicaciones, many=True)

        return Response(
            {"total": complicaciones.count(), "complicaciones": serializer.data},
        )

    @action(detail=False, methods=["get"])
    def requirieron_cirugia(self, _request):
        """Complicaciones que requirieron cirugía"""
        complicaciones = self.get_queryset().filter(requirio_cirugia=True)
        serializer = self.get_serializer(complicaciones, many=True)

        return Response(
            {"total": complicaciones.count(), "complicaciones": serializer.data},
        )

    @action(detail=False, methods=["get"])
    def estadisticas(self, _request):
        """Estadísticas generales de complicaciones"""
        queryset = self.get_queryset()

        return Response(
            {
                "total": queryset.count(),
                "por_severidad": list(
                    queryset.values("severidad").annotate(total=Count("id")),
                ),
                "por_tipo": list(
                    queryset.values("tipo_complicacion").annotate(total=Count("id")),
                ),
                "requirieron_cirugia": queryset.filter(requirio_cirugia=True).count(),
                "resolucion": list(
                    queryset.values("resolucion_complicacion").annotate(
                        total=Count("id"),
                    ),
                ),
            },
        )


# ═══════════════════════════════════════════════════════════════════════════
# APGAR SCORE DETALLADO VIEWSET - COMPLETO
# ═══════════════════════════════════════════════════════════════════════════


class ApgarScoreDetalladoViewSet(viewsets.ModelViewSet):
    """ViewSet COMPLETO para gestionar scores de Apgar detallados

    ACCIONES PERSONALIZADAS:
        pass
    - por_recien_nacido: Scores de un RN específico
    - scores_bajos: Scores menores a 7
    - scores_criticos: Scores menores a 4
    - estadisticas: Estadísticas de scores Apgar
    - comparativa_minutos: Comparación entre diferentes minutos
    """

    queryset = ApgarScoreDetallado.objects.all()
    serializer_class = ApgarScoreDetalladoSerializer
    permission_classes = [FetalMedicalPermission]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    filterset_fields = {
        "recien_nacido": ["exact"],
        "minuto_evaluacion": ["exact", "in"],
        "frecuencia_cardiaca": ["exact"],
        "esfuerzo_respiratorio": ["exact"],
    }

    ordering_fields = ["minuto_evaluacion", "fecha_registro"]
    ordering = ["minuto_evaluacion"]

    def get_queryset(self):
        """Optimiza queries"""
        return (
            super()
            .get_queryset()
            .select_related("recien_nacido", "recien_nacido__parto", "evaluador")
        )

    @action(detail=False, methods=["get"])
    def por_recien_nacido(self, request):
        """Scores de un recién nacido específico"""
        rn_id = request.query_params.get("recien_nacido_id")

        if not rn_id:
            return Response(
                {"error": "Debe proporcionar recien_nacido_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        scores = (
            self.get_queryset()
            .filter(recien_nacido_id=rn_id)
            .order_by("minuto_evaluacion")
        )

        serializer = self.get_serializer(scores, many=True)

        return Response({"total": scores.count(), "scores": serializer.data})

    @action(detail=False, methods=["get"])
    def scores_bajos(self, _request):
        """Scores Apgar menores a 7"""
        # Calculamos el score total en la query
        scores_bajos = []

        for score in self.get_queryset():
            if score.score_total < 7:
                scores_bajos.append(score)

        serializer = self.get_serializer(scores_bajos, many=True)

        return Response({"total": len(scores_bajos), "scores": serializer.data})

    @action(detail=False, methods=["get"])
    def scores_criticos(self, _request):
        """Scores Apgar menores a 4"""
        scores_criticos = []

        for score in self.get_queryset():
            if score.score_total < 4:
                scores_criticos.append(score)

        serializer = self.get_serializer(scores_criticos, many=True)

        return Response({"total": len(scores_criticos), "scores": serializer.data})

    @action(detail=False, methods=["get"])
    def estadisticas(self, _request):
        """Estadísticas generales de scores Apgar"""
        queryset = self.get_queryset()

        # Promedios por componente
        promedios = {
            "frecuencia_cardiaca": round(
                queryset.aggregate(Avg("frecuencia_cardiaca"))[
                    "frecuencia_cardiaca__avg"
                ]
                or 0,
                2,
            ),
            "esfuerzo_respiratorio": round(
                queryset.aggregate(Avg("esfuerzo_respiratorio"))[
                    "esfuerzo_respiratorio__avg"
                ]
                or 0,
                2,
            ),
            "tono_muscular": round(
                queryset.aggregate(Avg("tono_muscular"))["tono_muscular__avg"] or 0, 2,
            ),
            "irritabilidad_refleja": round(
                queryset.aggregate(Avg("irritabilidad_refleja"))[
                    "irritabilidad_refleja__avg"
                ]
                or 0,
                2,
            ),
            "coloracion": round(
                queryset.aggregate(Avg("coloracion"))["coloracion__avg"] or 0, 2,
            ),
        }

        return Response(
            {
                "total_evaluaciones": queryset.count(),
                "promedios_componentes": promedios,
                "por_minuto": list(
                    queryset.values("minuto_evaluacion").annotate(total=Count("id")),
                ),
            },
        )

    @action(detail=False, methods=["get"])
    def comparativa_minutos(self, request):
        """Comparación de scores entre diferentes minutos"""
        rn_id = request.query_params.get("recien_nacido_id")

        if not rn_id:
            return Response(
                {"error": "Debe proporcionar recien_nacido_id"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        scores = (
            self.get_queryset()
            .filter(recien_nacido_id=rn_id)
            .order_by("minuto_evaluacion")
        )

        comparativa = []
        for score in scores:
            comparativa.append(
                {
                    "minuto": score.minuto_evaluacion,
                    "score_total": score.score_total,
                    "desglose": score.get_desglose_componentes(),
                },
            )

        return Response({"comparativa": comparativa})
