"""=============================================================================
EXCEL GENERATOR - Generación de Reportes Excel con Gráficos
=============================================================================
Genera reportes Excel profesionales con múltiples hojas, gráficos y formato
Utiliza openpyxl para manipulación avanzada de Excel
=============================================================================
"""

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelGenerator:
    """Generador avanzado de archivos Excel con soporte para gráficos y formato
    """

    def __init__(self):
        """Init"""
        self.header_fill = PatternFill(
            start_color="3949AB", end_color="3949AB", fill_type="solid",
        )
        self.header_font = Font(color="FFFFFF", bold=True, size=12)
        self.title_font = Font(bold=True, size=16, color="1A237E")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_statistics_excel(
        self, stats_data: dict[str, Any], params: dict[str, Any],
    ) -> BytesIO:
        """Genera archivo Excel completo con estadísticas y gráficos

        Args:
            stats_data: Datos de estadísticas del sistema
            params: Parámetros del reporte (fechas, filtros, etc.)

        Returns:
            BytesIO con el contenido del archivo Excel

        """
        wb = Workbook()

        # Eliminar hoja por defecto
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Crear hojas
        self._create_summary_sheet(wb, stats_data, params)

        if "dashboard" in stats_data:
            self._create_dashboard_sheet(wb, stats_data["dashboard"])

        if "embarazos" in stats_data:
            self._create_pregnancies_sheet(wb, stats_data["embarazos"])

        if "controles" in stats_data:
            self._create_controls_sheet(wb, stats_data["controles"])

        if "partos" in stats_data:
            self._create_births_sheet(wb, stats_data["partos"])

        if "citas" in stats_data:
            self._create_appointments_sheet(wb, stats_data["citas"])

        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_summary_sheet(
        self, wb: Workbook, stats_data: dict[str, Any], params: dict[str, Any],
    ):
        """Crea hoja de resumen general"""
        ws = wb.create_sheet("Resumen General", 0)

        # Título
        ws["A1"] = "REPORTE DE ESTADÍSTICAS DEL SISTEMA"
        ws["A1"].font = self.title_font
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:D1")

        # Información del reporte
        row = 3
        ws[f"A{row}"] = "Fecha de Generación:"
        ws[f"B{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Período de Análisis:"
        ws[f"B{row}"] = (
            f"{params.get('fecha_inicio', 'N/A')} - {params.get('fecha_fin', 'N/A')}"
        )
        ws[f"A{row}"].font = Font(bold=True)

        # Resumen de módulos
        row += 3
        ws[f"A{row}"] = "RESUMEN POR MÓDULO"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws.merge_cells(f"A{row}:D{row}")

        row += 2
        headers = ["Módulo", "Total Registros", "Período Actual", "Observaciones"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # Datos de resumen
        row += 1
        modules_data = [
            (
                "Pacientes",
                stats_data.get("dashboard", {}).get("total_pacientes", 0),
                stats_data.get("dashboard", {}).get("nuevos_pacientes", 0),
                "Activos en sistema",
            ),
            (
                "Embarazos",
                stats_data.get("embarazos", {}).get("total_embarazos", 0),
                stats_data.get("dashboard", {}).get("embarazos_activos", 0),
                "Embarazos activos",
            ),
            (
                "Controles",
                stats_data.get("controles", {}).get("total_controles", 0),
                stats_data.get("dashboard", {}).get("controles_mes", 0),
                "Controles del mes",
            ),
            (
                "Partos",
                stats_data.get("partos", {}).get("total_partos", 0),
                stats_data.get("dashboard", {}).get("partos_mes", 0),
                "Partos del mes",
            ),
            (
                "Citas",
                stats_data.get("citas", {}).get("total_citas", 0),
                stats_data.get("dashboard", {}).get("citas_pendientes", 0),
                "Citas pendientes",
            ),
        ]

        for module_row in modules_data:
            for col, value in enumerate(module_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center" if col > 1 else "left")
            row += 1

        # Ajustar anchos de columna
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 25

    def _create_dashboard_sheet(self, wb: Workbook, dashboard_data: dict[str, Any]):
        """Crea hoja de dashboard con KPIs"""
        ws = wb.create_sheet("Dashboard")

        # Título
        ws["A1"] = "INDICADORES CLAVE (KPIs)"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Tabla de KPIs
        row = 3
        ws[f"A{row}"] = "Indicador"
        ws[f"B{row}"] = "Valor"
        ws[f"C{row}"] = "Descripción"

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Datos de KPIs
        kpis_data = [
            (
                "Total Pacientes",
                dashboard_data.get("total_pacientes", 0),
                "Pacientes registrados",
            ),
            (
                "Nuevos Pacientes",
                dashboard_data.get("nuevos_pacientes", 0),
                "Nuevos en el período",
            ),
            (
                "Embarazos Activos",
                dashboard_data.get("embarazos_activos", 0),
                "Embarazos en curso",
            ),
            (
                "Embarazos Alto Riesgo",
                dashboard_data.get("embarazos_alto_riesgo", 0),
                "Requieren atención",
            ),
            (
                "Controles del Mes",
                dashboard_data.get("controles_mes", 0),
                "Controles realizados",
            ),
            (
                "Citas Pendientes",
                dashboard_data.get("citas_pendientes", 0),
                "Por confirmar/realizar",
            ),
            ("Citas Hoy", dashboard_data.get("citas_hoy", 0), "Programadas para hoy"),
            (
                "Partos del Mes",
                dashboard_data.get("partos_mes", 0),
                "Partos realizados",
            ),
        ]

        row += 1
        for kpi_row in kpis_data:
            ws[f"A{row}"] = kpi_row[0]
            ws[f"B{row}"] = kpi_row[1]
            ws[f"C{row}"] = kpi_row[2]

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 2:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, size=11)
            row += 1

        # Gráfico de barras
        chart = BarChart()
        chart.title = "Indicadores Principales"
        chart.x_axis.title = "Indicadores"
        chart.y_axis.title = "Cantidad"

        data = Reference(ws, min_col=2, min_row=3, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "E3")

        # Ajustar anchos
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30

    def _create_pregnancies_sheet(self, wb: Workbook, pregnancy_data: dict[str, Any]):
        """Crea hoja de estadísticas de embarazos"""
        ws = wb.create_sheet("Embarazos")

        # Título
        ws["A1"] = "ESTADÍSTICAS DE EMBARAZOS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Datos generales
        row = 3
        ws[f"A{row}"] = "Total Embarazos:"
        ws[f"B{row}"] = pregnancy_data.get("total_embarazos", 0)
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Edad Promedio:"
        ws[f"B{row}"] = f"{pregnancy_data.get('edad_promedio', 0)} años"
        ws[f"A{row}"].font = Font(bold=True)

        # Distribución por riesgo
        row += 3
        ws[f"A{row}"] = "DISTRIBUCIÓN POR NIVEL DE RIESGO"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:C{row}")

        row += 2
        ws[f"A{row}"] = "Nivel de Riesgo"
        ws[f"B{row}"] = "Cantidad"
        ws[f"C{row}"] = "Porcentaje"

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        # Datos por riesgo
        por_riesgo = pregnancy_data.get("por_riesgo", [])
        total_embarazos = pregnancy_data.get(
            "total_embarazos", 1,
        )  # Evitar división por cero

        row += 1
        start_data_row = row
        for riesgo_item in por_riesgo:
            clasificacion = riesgo_item.get("clasificacion_riesgo", "N/A")
            total = riesgo_item.get("total", 0)
            porcentaje = (total / total_embarazos * 100) if total_embarazos > 0 else 0

            ws[f"A{row}"] = clasificacion.title()
            ws[f"B{row}"] = total
            ws[f"C{row}"] = f"{porcentaje:.1f}%"

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col > 1:
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        # Gráfico de pastel
        if por_riesgo:
            pie = PieChart()
            pie.title = "Distribución por Nivel de Riesgo"

            labels = Reference(ws, min_col=1, min_row=start_data_row, max_row=row - 1)
            data = Reference(ws, min_col=2, min_row=start_data_row - 1, max_row=row - 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)

            ws.add_chart(pie, "E3")

        # Ajustar anchos
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_controls_sheet(self, wb: Workbook, controls_data: dict[str, Any]):
        """Crea hoja de estadísticas de controles"""
        ws = wb.create_sheet("Controles Prenatales")

        # Título
        ws["A1"] = "ESTADÍSTICAS DE CONTROLES PRENATALES"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Datos generales
        row = 3
        ws[f"A{row}"] = "Total Controles:"
        ws[f"B{row}"] = controls_data.get("total_controles", 0)
        ws[f"A{row}"].font = Font(bold=True)

        # Promedios clínicos
        row += 3
        ws[f"A{row}"] = "PROMEDIOS CLÍNICOS"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:B{row}")

        row += 2
        ws[f"A{row}"] = "Parámetro"
        ws[f"B{row}"] = "Valor Promedio"

        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        promedios = controls_data.get("promedios", {})

        row += 1
        clinical_data = [
            ("Peso", f"{promedios.get('peso_promedio', 0):.1f} kg"),
            (
                "Presión Sistólica",
                f"{promedios.get('presion_sistolica_promedio', 0):.0f} mmHg",
            ),
            (
                "Presión Diastólica",
                f"{promedios.get('presion_diastolica_promedio', 0):.0f} mmHg",
            ),
        ]

        for param, value in clinical_data:
            ws[f"A{row}"] = param
            ws[f"B{row}"] = value

            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 2:
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        # Ajustar anchos
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_births_sheet(self, wb: Workbook, births_data: dict[str, Any]):
        """Crea hoja de estadísticas de partos"""
        ws = wb.create_sheet("Partos")

        # Título
        ws["A1"] = "ESTADÍSTICAS DE PARTOS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Datos generales
        row = 3
        ws[f"A{row}"] = "Total Partos:"
        ws[f"B{row}"] = births_data.get("total_partos", 0)
        ws[f"A{row}"].font = Font(bold=True)

        # Distribución por tipo
        row += 3
        ws[f"A{row}"] = "DISTRIBUCIÓN POR TIPO DE PARTO"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{row}:C{row}")

        row += 2
        ws[f"A{row}"] = "Tipo de Parto"
        ws[f"B{row}"] = "Cantidad"
        ws[f"C{row}"] = "Porcentaje"

        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border

        por_tipo = births_data.get("por_tipo", [])
        total_partos = births_data.get("total_partos", 1)

        row += 1
        for tipo_item in por_tipo:
            tipo = tipo_item.get("tipo_parto", "N/A")
            total = tipo_item.get("total", 0)
            porcentaje = (total / total_partos * 100) if total_partos > 0 else 0

            ws[f"A{row}"] = tipo.replace("_", " ").title()
            ws[f"B{row}"] = total
            ws[f"C{row}"] = f"{porcentaje:.1f}%"

            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col > 1:
                    cell.alignment = Alignment(horizontal="center")
            row += 1

        # Ajustar anchos
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_appointments_sheet(
        self, wb: Workbook, appointments_data: dict[str, Any],
    ):
        """Crea hoja de estadísticas de citas"""
        ws = wb.create_sheet("Citas")

        # Título
        ws["A1"] = "ESTADÍSTICAS DE CITAS"
        ws["A1"].font = self.title_font
        ws.merge_cells("A1:C1")

        # Datos generales
        row = 3
        data_items = [
            ("Total Citas", appointments_data.get("total_citas", 0)),
            ("Completadas", appointments_data.get("completadas", 0)),
            ("Canceladas", appointments_data.get("canceladas", 0)),
            ("No Asistió", appointments_data.get("no_asistidas", 0)),
            (
                "Tasa de Asistencia",
                f"{appointments_data.get('tasa_asistencia', 0):.1f}%",
            ),
            (
                "Tasa de Cancelación",
                f"{appointments_data.get('tasa_cancelacion', 0):.1f}%",
            ),
        ]

        for label, value in data_items:
            ws[f"A{row}"] = label + ":"
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Ajustar anchos
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20


# Instancia global del generador
excel_generator = ExcelGenerator()
